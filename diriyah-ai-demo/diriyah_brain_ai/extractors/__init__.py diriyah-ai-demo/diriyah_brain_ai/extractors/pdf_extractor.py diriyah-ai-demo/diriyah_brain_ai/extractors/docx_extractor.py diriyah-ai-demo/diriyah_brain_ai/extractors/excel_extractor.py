"""Excel spreadsheet extraction utilities.

This module converts Microsoft Excel workbooks into a plain text
representation by concatenating the values from every sheet.  It
leverages the ``openpyxl`` library to parse the workbook in memory.
If parsing fails, an empty string is returned.
"""

from __future__ import annotations

import io
from typing import Optional

# The openpyxl library is optional.  It is imported lazily inside the
# extraction function.  If unavailable, the extractor simply returns
# an empty string instead of raising.


def extract_excel_bytes(data: bytes) -> str:
    """Extract plain text from an Excel workbook represented by raw bytes.

    Args:
        data: Raw XLSX data.

    Returns:
        A newline separated string containing all cell values from all
        sheets.  If ``openpyxl`` is not installed or the workbook
        cannot be loaded, returns an empty string.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return ""
    try:
        buf = io.BytesIO(data)
        wb = load_workbook(buf, data_only=True)
        parts = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    parts.append(str(cell))
        return "\n".join(parts)
    except Exception:
        return ""
