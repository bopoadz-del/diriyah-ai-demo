from openpyxl import load_workbook

def extract_excel_cells(path:str,max_cells=200)->str:
  try:
    wb=load_workbook(filename=path,read_only=True,data_only=True); ws=wb.active; out=[]
    for row in ws.iter_rows(values_only=True):
      if len(out)>=max_cells: break
      out.append(', '.join('' if c is None else str(c) for c in row))
    return '\n'.join(out)
  except Exception as e:
    return f'Excel extraction error: {e}'
