from __future__ import annotations

import base64
import json
import logging
import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from enum import Enum
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query
from fastapi.responses import FileResponse

try:  # pragma: no cover - optional templating dependency
    from jinja2 import Environment, FileSystemLoader
except Exception as exc:  # pragma: no cover - diagnostics only
    Environment = None  # type: ignore[assignment]
    FileSystemLoader = None  # type: ignore[assignment]
    _jinja_import_error: Optional[Exception] = exc
else:  # pragma: no cover
    _jinja_import_error = None

logger = logging.getLogger(__name__)

plt = None  # type: ignore[assignment]
_matplotlib_import_error: Optional[Exception] = None

try:  # pragma: no cover - optional styling
    import seaborn as sns  # type: ignore
except Exception as exc:  # pragma: no cover - diagnostics only
    sns = None  # type: ignore[assignment]
    _seaborn_import_error: Optional[Exception] = exc
else:  # pragma: no cover
    _seaborn_import_error = None


def _load_matplotlib() -> None:
    global plt, _matplotlib_import_error
    if plt is not None or _matplotlib_import_error is not None:
        return
    try:  # pragma: no cover - optional dependency for Render deployments
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as matplotlib_pyplot  # type: ignore
    except Exception as exc:  # pragma: no cover - diagnostics only
        plt = None  # type: ignore[assignment]
        _matplotlib_import_error = exc
    else:  # pragma: no cover - plotting available
        plt = matplotlib_pyplot  # type: ignore[assignment]
        _matplotlib_import_error = None

try:  # pragma: no cover - optional rich charts
    import plotly.graph_objects as go  # type: ignore
except Exception as exc:  # pragma: no cover - diagnostics only
    go = None  # type: ignore[assignment]
    _plotly_import_error: Optional[Exception] = exc
else:  # pragma: no cover
    _plotly_import_error = None

try:  # pragma: no cover - optional PDF rendering
    from weasyprint import HTML  # type: ignore
except Exception as exc:  # pragma: no cover - diagnostics only
    HTML = None  # type: ignore[assignment]
    _weasyprint_import_error: Optional[Exception] = exc
else:  # pragma: no cover
    _weasyprint_import_error = None

try:  # pragma: no cover - optional Excel support
    from openpyxl import Workbook  # type: ignore
except Exception as exc:  # pragma: no cover - diagnostics only
    Workbook = None  # type: ignore[assignment]
    _openpyxl_import_error: Optional[Exception] = exc
else:  # pragma: no cover
    _openpyxl_import_error = None

try:  # pragma: no cover - optional AI insights
    from openai import OpenAI  # type: ignore
except Exception as exc:  # pragma: no cover - diagnostics only
    OpenAI = None  # type: ignore[assignment]
    _openai_import_error: Optional[Exception] = exc
else:  # pragma: no cover
    _openai_import_error = None


def _optional_db() -> Iterable[Any]:
    """Yield an optional database session when available."""

    try:
        from backend.backend.db import get_db as real_get_db  # local import for resilience
    except Exception:  # pragma: no cover - DB optional during debugging
        yield None
        return

    try:  # pragma: no cover - relies on runtime environment
        yield from real_get_db()
    except Exception as exc:  # pragma: no cover - diagnostics only
        logger.warning("Database session unavailable: %s", exc)
        yield None


@dataclass
class UserContext:
    """Lightweight user context used for Render debugging deployments."""

    id: str
    name: str
    email: str = "render-debug@diriyah.ai"


def get_current_user() -> UserContext:
    """Return a deterministic user context for analytics stubs."""

    return UserContext(id="user-001", name="Render Debug User")


class ReportType(Enum):
    DAILY_SUMMARY = "daily_summary"
    WEEKLY_PROGRESS = "weekly_progress"
    MONTHLY_OVERVIEW = "monthly_overview"
    PROJECT_STATUS = "project_status"
    SAFETY_REPORT = "safety_report"
    QUALITY_REPORT = "quality_report"
    FINANCIAL_REPORT = "financial_report"
    EXECUTIVE_SUMMARY = "executive_summary"
    CUSTOM = "custom"


class ReportFormat(Enum):
    PDF = "pdf"
    HTML = "html"
    EXCEL = "excel"
    JSON = "json"


@dataclass
class TableData:
    """Simple table abstraction that does not rely on pandas."""

    headers: List[str]
    rows: List[Dict[str, Any]]

    def to_html(self) -> str:
        thead = "".join(f"<th>{header}</th>" for header in self.headers)
        body_rows = []
        for row in self.rows:
            cells = "".join(f"<td>{row.get(header, '')}</td>" for header in self.headers)
            body_rows.append(f"<tr>{cells}</tr>")
        tbody = "".join(body_rows)
        return f"<table class='data-table'><thead><tr>{thead}</tr></thead><tbody>{tbody}</tbody></table>"


@dataclass
class ReportData:
    title: str
    period: str
    generated_at: datetime
    project_id: Optional[str]
    metrics: Dict[str, Any] = field(default_factory=dict)
    charts: Dict[str, str] = field(default_factory=dict)
    tables: Dict[str, TableData] = field(default_factory=dict)
    insights: List[str] = field(default_factory=list)
    recommendations: List[str] = field(default_factory=list)
    summary: str = ""


@dataclass
class ProjectRecord:
    id: str
    user_id: str
    name: str
    status: str
    progress: float
    created_at: datetime


@dataclass
class TaskRecord:
    id: str
    user_id: str
    project_id: str
    title: str
    priority: str
    status: str
    created_at: datetime
    due_date: Optional[datetime]
    completed: bool


@dataclass
class AlertRecord:
    id: str
    user_id: str
    project_id: str
    title: str
    severity: str
    created_at: datetime


_NOW = datetime.utcnow()
_SAMPLE_PROJECTS: List[ProjectRecord] = [
    ProjectRecord(
        id="PRJ-001",
        user_id="user-001",
        name="Cultural District Core",
        status="active",
        progress=72.5,
        created_at=_NOW - timedelta(days=120),
    ),
    ProjectRecord(
        id="PRJ-002",
        user_id="user-001",
        name="Transit Hub Phase 2",
        status="planning",
        progress=18.0,
        created_at=_NOW - timedelta(days=45),
    ),
    ProjectRecord(
        id="PRJ-003",
        user_id="user-002",
        name="Logistics Spine",
        status="active",
        progress=54.0,
        created_at=_NOW - timedelta(days=80),
    ),
]

_SAMPLE_TASKS: List[TaskRecord] = [
    TaskRecord(
        id="TSK-001",
        user_id="user-001",
        project_id="PRJ-001",
        title="Pour level 12 core",
        priority="high",
        status="completed",
        created_at=_NOW - timedelta(days=6),
        due_date=_NOW - timedelta(days=3),
        completed=True,
    ),
    TaskRecord(
        id="TSK-002",
        user_id="user-001",
        project_id="PRJ-001",
        title="Install façade mockup",
        priority="medium",
        status="in_progress",
        created_at=_NOW - timedelta(days=4),
        due_date=_NOW + timedelta(days=5),
        completed=False,
    ),
    TaskRecord(
        id="TSK-003",
        user_id="user-001",
        project_id="PRJ-002",
        title="Finalize transport permits",
        priority="high",
        status="completed",
        created_at=_NOW - timedelta(days=10),
        due_date=_NOW - timedelta(days=2),
        completed=True,
    ),
    TaskRecord(
        id="TSK-004",
        user_id="user-001",
        project_id="PRJ-002",
        title="Award excavation package",
        priority="low",
        status="pending",
        created_at=_NOW - timedelta(days=1),
        due_date=_NOW + timedelta(days=14),
        completed=False,
    ),
    TaskRecord(
        id="TSK-005",
        user_id="user-002",
        project_id="PRJ-003",
        title="Commission logistics yard",
        priority="medium",
        status="in_progress",
        created_at=_NOW - timedelta(days=2),
        due_date=_NOW + timedelta(days=7),
        completed=False,
    ),
]

_SAMPLE_ALERTS: List[AlertRecord] = [
    AlertRecord(
        id="ALT-001",
        user_id="user-001",
        project_id="PRJ-001",
        title="Tower crane wind alarm",
        severity="critical",
        created_at=_NOW - timedelta(days=2, hours=5),
    ),
    AlertRecord(
        id="ALT-002",
        user_id="user-001",
        project_id="PRJ-001",
        title="Concrete strength test pending",
        severity="medium",
        created_at=_NOW - timedelta(days=3),
    ),
    AlertRecord(
        id="ALT-003",
        user_id="user-001",
        project_id="PRJ-002",
        title="Utility diversion coordination",
        severity="high",
        created_at=_NOW - timedelta(days=1, hours=3),
    ),
    AlertRecord(
        id="ALT-004",
        user_id="user-002",
        project_id="PRJ-003",
        title="Logistics gate access",
        severity="medium",
        created_at=_NOW - timedelta(days=1),
    ),
]


class AutomatedReportGenerator:
    """Composable report generator with Render-friendly fallbacks."""

    def __init__(
        self,
        db: Any = None,
        openai_api_key: Optional[str] = None,
        template_dir: str = "./templates",
    ) -> None:
        self.db = db
        if openai_api_key and OpenAI is not None:
            self.openai_client = OpenAI(api_key=openai_api_key)
        else:
            self.openai_client = None
            if openai_api_key and _openai_import_error:
                logger.warning("OpenAI SDK unavailable: %s", _openai_import_error)

        self.template_dir = Path(template_dir)
        self.template_dir.mkdir(exist_ok=True, parents=True)
        self.jinja_env: Optional[Environment]
        if Environment is not None and FileSystemLoader is not None:
            self.jinja_env = Environment(loader=FileSystemLoader(str(self.template_dir)))
            self._create_default_templates()
        else:
            self.jinja_env = None
            if _jinja_import_error:
                logger.warning("Jinja2 unavailable: %s", _jinja_import_error)

        self.output_dir = Path("./reports")
        self.output_dir.mkdir(exist_ok=True, parents=True)

        _load_matplotlib()
        if plt is not None and sns is not None:  # pragma: no cover - styling only
            try:
                plt.style.use("seaborn-v0_8-darkgrid")
                sns.set_palette("husl")
            except Exception as exc:  # pragma: no cover - diagnostics
                logger.debug("Failed to apply seaborn styling: %s", exc)

        if _matplotlib_import_error:
            logger.warning("Matplotlib unavailable: %s", _matplotlib_import_error)
        if _weasyprint_import_error:
            logger.warning("WeasyPrint unavailable: %s", _weasyprint_import_error)
        if _openpyxl_import_error:
            logger.debug("openpyxl unavailable: %s", _openpyxl_import_error)

    async def generate_report(
        self,
        report_type: ReportType,
        report_format: ReportFormat,
        user_id: str,
        project_id: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        custom_params: Optional[Dict[str, Any]] = None,
    ) -> str:
        if date_to is None:
            date_to = datetime.utcnow()
        if date_from is None:
            if report_type == ReportType.DAILY_SUMMARY:
                date_from = date_to - timedelta(days=1)
            elif report_type == ReportType.WEEKLY_PROGRESS:
                date_from = date_to - timedelta(days=7)
            elif report_type == ReportType.MONTHLY_OVERVIEW:
                date_from = date_to - timedelta(days=30)
            else:
                date_from = date_to - timedelta(days=14)

        report_data = await self._collect_report_data(
            report_type=report_type,
            user_id=user_id,
            project_id=project_id,
            date_from=date_from,
            date_to=date_to,
            custom_params=custom_params,
        )

        charts = await self._generate_charts(report_data)
        report_data.charts = charts

        if self.openai_client is not None:
            insights = await self._generate_ai_insights(report_data)
            report_data.insights = insights

        if report_format == ReportFormat.PDF:
            return await self._generate_pdf_report(report_data, report_type)
        if report_format == ReportFormat.HTML:
            return await self._generate_html_report(report_data, report_type)
        if report_format == ReportFormat.EXCEL:
            return await self._generate_excel_report(report_data, report_type)
        if report_format == ReportFormat.JSON:
            return await self._generate_json_report(report_data, report_type)
        raise ValueError(f"Unsupported report format: {report_format}")

    async def _collect_report_data(
        self,
        report_type: ReportType,
        user_id: str,
        project_id: Optional[str],
        date_from: datetime,
        date_to: datetime,
        custom_params: Optional[Dict[str, Any]],
    ) -> ReportData:
        projects = [
            project
            for project in _SAMPLE_PROJECTS
            if project.user_id == user_id and (project_id is None or project.id == project_id)
        ]
        tasks = [
            task
            for task in _SAMPLE_TASKS
            if task.user_id == user_id
            and date_from <= task.created_at <= date_to
            and (project_id is None or task.project_id == project_id)
        ]
        alerts = [
            alert
            for alert in _SAMPLE_ALERTS
            if alert.user_id == user_id
            and date_from <= alert.created_at <= date_to
            and (project_id is None or alert.project_id == project_id)
        ]

        metrics: Dict[str, Any] = {}
        metrics["total_projects"] = len(projects)
        metrics["active_projects"] = sum(1 for project in projects if project.status == "active")
        metrics["avg_progress"] = round(
            sum(project.progress for project in projects) / len(projects), 2
        ) if projects else 0.0

        metrics["total_tasks"] = len(tasks)
        metrics["completed_tasks"] = sum(1 for task in tasks if task.completed)
        metrics["pending_tasks"] = metrics["total_tasks"] - metrics["completed_tasks"]
        metrics["completion_rate"] = (
            round(metrics["completed_tasks"] / metrics["total_tasks"] * 100, 2)
            if metrics["total_tasks"]
            else 0.0
        )

        metrics["total_alerts"] = len(alerts)
        metrics["critical_alerts"] = sum(1 for alert in alerts if alert.severity == "critical")
        metrics["high_alerts"] = sum(1 for alert in alerts if alert.severity == "high")
        metrics["medium_alerts"] = sum(1 for alert in alerts if alert.severity == "medium")

        tables: Dict[str, TableData] = {}
        if projects:
            tables["projects"] = TableData(
                headers=["Name", "Status", "Progress", "Created"],
                rows=[
                    {
                        "Name": project.name,
                        "Status": project.status.title(),
                        "Progress": f"{project.progress:.1f}%",
                        "Created": project.created_at.strftime("%Y-%m-%d"),
                    }
                    for project in projects
                ],
            )
        if tasks:
            tables["tasks"] = TableData(
                headers=["Title", "Priority", "Status", "Due Date"],
                rows=[
                    {
                        "Title": task.title,
                        "Priority": task.priority.title(),
                        "Status": task.status.replace("_", " ").title(),
                        "Due Date": task.due_date.strftime("%Y-%m-%d") if task.due_date else "N/A",
                    }
                    for task in tasks
                ],
            )
        if alerts:
            tables["alerts"] = TableData(
                headers=["Title", "Severity", "Date"],
                rows=[
                    {
                        "Title": alert.title,
                        "Severity": alert.severity.title(),
                        "Date": alert.created_at.strftime("%Y-%m-%d %H:%M"),
                    }
                    for alert in sorted(alerts, key=lambda item: item.created_at, reverse=True)[:10]
                ],
            )

        summary = self._generate_summary(metrics)
        period_str = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"

        return ReportData(
            title=self._get_report_title(report_type),
            period=period_str,
            generated_at=datetime.utcnow(),
            project_id=project_id,
            metrics=metrics,
            tables=tables,
            summary=summary,
        )
    async def _generate_charts(self, report_data: ReportData) -> Dict[str, str]:
        charts: Dict[str, str] = {}
        _load_matplotlib()
        if plt is None:
            return charts

        try:
            fig, ax = plt.subplots(figsize=(8, 5))
            labels = [
                "Total\nProjects",
                "Active\nProjects",
                "Total\nTasks",
                "Completed\nTasks",
                "Pending\nTasks",
            ]
            values = [
                report_data.metrics.get("total_projects", 0),
                report_data.metrics.get("active_projects", 0),
                report_data.metrics.get("total_tasks", 0),
                report_data.metrics.get("completed_tasks", 0),
                report_data.metrics.get("pending_tasks", 0),
            ]
            bars = ax.bar(labels, values, color=["#3B82F6", "#10B981", "#F59E0B", "#8B5CF6", "#EF4444"])
            for bar in bars:
                height = bar.get_height()
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    height,
                    f"{int(height)}",
                    ha="center",
                    va="bottom",
                    fontweight="bold",
                )
            ax.set_ylabel("Count")
            ax.set_title("Project & Task Overview")
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            fig.tight_layout()
            charts["overview"] = self._fig_to_base64(fig)
            plt.close(fig)
        except Exception as exc:  # pragma: no cover - diagnostics only
            logger.debug("Failed to build overview chart: %s", exc)

        if go is not None:
            try:
                completion_rate = report_data.metrics.get("completion_rate", 0.0)
                fig = go.Figure(
                    go.Indicator(
                        mode="gauge+number",
                        value=completion_rate,
                        title={"text": "Task Completion"},
                        gauge={
                            "axis": {"range": [0, 100]},
                            "bar": {"color": "#10B981"},
                            "steps": [
                                {"range": [0, 50], "color": "#fee2e2"},
                                {"range": [50, 80], "color": "#fef3c7"},
                                {"range": [80, 100], "color": "#d1fae5"},
                            ],
                        },
                    )
                )
                charts["completion_gauge"] = self._plotly_to_base64(fig)
            except Exception as exc:  # pragma: no cover - diagnostics only
                logger.debug("Failed to build gauge chart: %s", exc)

        if plt is not None and "projects" in report_data.tables:
            try:
                table = report_data.tables["projects"]
                fig, ax = plt.subplots(figsize=(8, 5))
                names = [row["Name"] for row in table.rows]
                progress = [float(row["Progress"].rstrip("%")) for row in table.rows]
                bars = ax.barh(names, progress, color="#3B82F6")
                for bar, value in zip(bars, progress):
                    ax.text(value + 2, bar.get_y() + bar.get_height() / 2, f"{value:.1f}%", va="center")
                ax.set_xlabel("Progress (%)")
                ax.set_title("Project Progress")
                ax.set_xlim(0, 110)
                ax.spines["top"].set_visible(False)
                ax.spines["right"].set_visible(False)
                fig.tight_layout()
                charts["project_progress"] = self._fig_to_base64(fig)
                plt.close(fig)
            except Exception as exc:  # pragma: no cover - diagnostics only
                logger.debug("Failed to build project progress chart: %s", exc)

        if plt is not None and report_data.metrics.get("total_alerts", 0) > 0:
            try:
                fig, ax = plt.subplots(figsize=(6, 6))
                severities = ["Critical", "High", "Medium"]
                counts = [
                    report_data.metrics.get("critical_alerts", 0),
                    report_data.metrics.get("high_alerts", 0),
                    report_data.metrics.get("medium_alerts", 0),
                ]
                data = [(label, count) for label, count in zip(severities, counts) if count > 0]
                if data:
                    labels, values = zip(*data)
                    ax.pie(values, labels=labels, autopct="%1.1f%%", startangle=90)
                    ax.set_title("Alert Severity Distribution")
                    fig.tight_layout()
                    charts["alerts_pie"] = self._fig_to_base64(fig)
                plt.close(fig)
            except Exception as exc:  # pragma: no cover - diagnostics only
                logger.debug("Failed to build alerts pie chart: %s", exc)

        return charts

    async def _generate_ai_insights(self, report_data: ReportData) -> List[str]:
        if self.openai_client is None:
            return [
                "AI insights unavailable in Render debug mode.",
                f"Completion rate stands at {report_data.metrics.get('completion_rate', 0.0):.1f}%.",
                f"Active alerts recorded: {report_data.metrics.get('total_alerts', 0)}.",
            ]

        try:  # pragma: no cover - requires network
            summary_context = json.dumps(report_data.metrics, indent=2)
            response = self.openai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are an expert construction analyst. Provide concise,"
                            " actionable observations."
                        ),
                    },
                    {
                        "role": "user",
                        "content": (
                            "Review these metrics and provide 5 insights:\n" f"{summary_context}"
                        ),
                    },
                ],
                max_tokens=300,
            )
            ai_text = response.choices[0].message.content
            lines = [line.strip() for line in ai_text.splitlines() if line.strip()]
            insights: List[str] = []
            for line in lines:
                cleaned = line.lstrip("0123456789.-) ")
                if cleaned:
                    insights.append(cleaned)
            return insights[:5] or [
                "Projects continue progressing steadily.",
                "Alert volume remains manageable.",
            ]
        except Exception as exc:  # pragma: no cover - diagnostics only
            logger.warning("Failed to generate AI insights: %s", exc)
            return [
                "Unable to retrieve AI-generated insights at this time.",
                f"Completion rate: {report_data.metrics.get('completion_rate', 0.0):.1f}%.",
            ]

    async def _generate_pdf_report(self, report_data: ReportData, report_type: ReportType) -> str:
        html_path = await self._generate_html_report(report_data, report_type)
        if HTML is None:
            return html_path
        pdf_filename = f"report_{report_type.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        pdf_path = self.output_dir / pdf_filename
        try:
            HTML(html_path).write_pdf(pdf_path)
            return str(pdf_path)
        except Exception as exc:  # pragma: no cover - diagnostics only
            logger.warning("Failed to write PDF, returning HTML instead: %s", exc)
            return html_path

    async def _generate_html_report(self, report_data: ReportData, report_type: ReportType) -> str:
        template_name = f"{report_type.value}_template.html"
        if self.jinja_env is None:
            html_content = self._render_basic_html(report_data)
        else:
            try:
                template = self.jinja_env.get_template(template_name)
            except Exception:
                template = self.jinja_env.get_template("default_template.html")
            html_content = template.render(
                report=report_data,
                generated_date=datetime.now().strftime("%B %d, %Y at %H:%M"),
                **report_data.metrics,
            )
        html_filename = f"report_{report_type.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        html_path = self.output_dir / html_filename
        html_path.write_text(html_content, encoding="utf-8")
        return str(html_path)

    def _render_basic_html(self, report_data: ReportData) -> str:
        """Render a lightweight HTML report when Jinja2 is unavailable."""

        metrics_html = "".join(
            f"<li><strong>{key.replace('_', ' ').title()}:</strong> {value}</li>"
            for key, value in report_data.metrics.items()
        )
        insights_html = "".join(f"<li>{insight}</li>" for insight in report_data.insights) or "<li>No insights available.</li>"
        recommendations_html = "".join(
            f"<li>{recommendation}</li>" for recommendation in report_data.recommendations
        ) or "<li>No recommendations available.</li>"

        tables_sections: List[str] = []
        for table_name, table in report_data.tables.items():
            header_html = "".join(f"<th>{column}</th>" for column in table.headers)
            rows_html = "".join(
                "<tr>" + "".join(f"<td>{row.get(column, '')}</td>" for column in table.headers) + "</tr>"
                for row in table.rows
            )
            tables_sections.append(
                f"<section><h3>{table_name.title()} Details</h3><table><thead><tr>{header_html}</tr></thead><tbody>{rows_html}</tbody></table></section>"
            )

        tables_html = "".join(tables_sections) or "<p>No tabular data available.</p>"

        return f"""
<!DOCTYPE html>
<html lang=\"en\">
  <head>
    <meta charset=\"utf-8\">
    <title>{report_data.title}</title>
    <style>
      body {{ font-family: Arial, sans-serif; background: #F9FAFB; color: #1F2937; margin: 0; padding: 24px; }}
      h1 {{ color: #92400E; }}
      section {{ background: #FFFFFF; border-radius: 12px; padding: 16px; margin-bottom: 16px; box-shadow: 0 1px 4px rgba(0,0,0,0.06); }}
      ul {{ padding-left: 20px; }}
      table {{ width: 100%; border-collapse: collapse; }}
      th, td {{ border: 1px solid #E5E7EB; padding: 8px; text-align: left; }}
      th {{ background: #F3F4F6; }}
    </style>
  </head>
  <body>
    <h1>{report_data.title}</h1>
    <p><strong>Period:</strong> {report_data.period}</p>
    <p><strong>Generated:</strong> {datetime.now().strftime('%B %d, %Y at %H:%M')}</p>
    <section>
      <h2>Metrics</h2>
      <ul>{metrics_html}</ul>
    </section>
    <section>
      <h2>Key Insights</h2>
      <ul>{insights_html}</ul>
    </section>
    <section>
      <h2>Recommendations</h2>
      <ul>{recommendations_html}</ul>
    </section>
    {tables_html}
  </body>
</html>
"""

    async def _generate_excel_report(self, report_data: ReportData, report_type: ReportType) -> str:
        if Workbook is None:
            logger.warning("openpyxl unavailable, returning JSON instead of Excel")
            return await self._generate_json_report(report_data, report_type)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Summary"
        sheet.append(["Metric", "Value"])
        for key, value in report_data.metrics.items():
            sheet.append([key, value])
        for table_name, table in report_data.tables.items():
            ws = workbook.create_sheet(title=table_name.title()[:31])
            ws.append(table.headers)
            for row in table.rows:
                ws.append([row.get(header, "") for header in table.headers])
        if report_data.insights:
            ws = workbook.create_sheet(title="Insights")
            ws.append(["Insight"])
            for insight in report_data.insights:
                ws.append([insight])
        if report_data.recommendations:
            ws = workbook.create_sheet(title="Recommendations")
            ws.append(["Recommendation"])
            for recommendation in report_data.recommendations:
                ws.append([recommendation])
        excel_filename = f"report_{report_type.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_path = self.output_dir / excel_filename
        workbook.save(excel_path)
        return str(excel_path)

    async def _generate_json_report(self, report_data: ReportData, report_type: ReportType) -> str:
        json_filename = f"report_{report_type.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        json_path = self.output_dir / json_filename
        payload = {
            "title": report_data.title,
            "period": report_data.period,
            "generated_at": report_data.generated_at.isoformat(),
            "project_id": report_data.project_id,
            "metrics": report_data.metrics,
            "insights": report_data.insights,
            "recommendations": report_data.recommendations,
            "summary": report_data.summary,
            "tables": {
                name: table.rows
                for name, table in report_data.tables.items()
            },
        }
        json_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        return str(json_path)

    def _create_default_templates(self) -> None:
        template_path = self.template_dir / "default_template.html"
        if template_path.exists():
            return
        template_html = """
<!DOCTYPE html>
<html>
<head>
    <meta charset=\"UTF-8\">
    <title>{{ report.title }}</title>
    <style>
        body { font-family: Arial, sans-serif; background: #F9FAFB; color: #1F2937; margin: 0; }
        .container { max-width: 960px; margin: 0 auto; padding: 32px; }
        .header { background: linear-gradient(135deg, #D97706, #92400E); color: #fff; padding: 32px; border-radius: 16px; }
        .header h1 { margin: 0 0 8px 0; font-size: 32px; }
        .metrics { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin: 32px 0; }
        .metric-card { background: #fff; padding: 20px; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.05); border-left: 4px solid #D97706; }
        .metric-card .value { font-size: 28px; font-weight: bold; color: #D97706; }
        .metric-card .label { font-size: 12px; text-transform: uppercase; letter-spacing: 1px; color: #6B7280; }
        .section { background: #fff; padding: 24px; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.05); margin-bottom: 24px; }
        .section h2 { margin-top: 0; font-size: 24px; border-bottom: 2px solid #D97706; padding-bottom: 8px; }
        .chart-container { text-align: center; margin: 16px 0; }
        .chart-container img { max-width: 100%; border-radius: 12px; }
        .insights-list, .recommendations-list { list-style: none; padding: 0; }
        .insights-list li { background: #FFFBEB; border-left: 4px solid #D97706; padding: 12px; margin-bottom: 8px; border-radius: 4px; }
        .recommendations-list li { background: #ECFDF5; border-left: 4px solid #10B981; padding: 12px; margin-bottom: 8px; border-radius: 4px; }
        table.data-table { width: 100%; border-collapse: collapse; margin-top: 16px; }
        table.data-table th, table.data-table td { border-bottom: 1px solid #E5E7EB; padding: 12px; text-align: left; }
        table.data-table th { background: #F3F4F6; font-weight: 600; }
        .footer { text-align: center; color: #6B7280; font-size: 12px; margin-top: 32px; }
    </style>
</head>
<body>
    <div class=\"container\">
        <div class=\"header\">
            <h1>{{ report.title }}</h1>
            <p>Period: {{ report.period }}</p>
            <p>Generated: {{ generated_date }}</p>
        </div>
        <div class=\"metrics\">
            <div class=\"metric-card\"><div class=\"value\">{{ total_projects }}</div><div class=\"label\">Total Projects</div></div>
            <div class=\"metric-card\"><div class=\"value\">{{ active_projects }}</div><div class=\"label\">Active Projects</div></div>
            <div class=\"metric-card\"><div class=\"value\">{{ '%.1f'|format(completion_rate) }}%</div><div class=\"label\">Completion Rate</div></div>
            <div class=\"metric-card\"><div class=\"value\">{{ total_alerts }}</div><div class=\"label\">Total Alerts</div></div>
        </div>
        <div class=\"section\">
            <h2>Executive Summary</h2>
            <p>{{ report.summary }}</p>
        </div>
        {% if report.charts %}
        <div class=\"section\">
            <h2>Visual Analytics</h2>
            {% for chart_name, chart_data in report.charts.items() %}
            <div class=\"chart-container\">
                <img src=\"data:image/png;base64,{{ chart_data }}\" alt=\"{{ chart_name }}\">
            </div>
            {% endfor %}
        </div>
        {% endif %}
        {% if report.insights %}
        <div class=\"section\">
            <h2>Key Insights</h2>
            <ul class=\"insights-list\">
                {% for insight in report.insights %}
                <li>{{ insight }}</li>
                {% endfor %}
            </ul>
        </div>
        {% endif %}
        {% if report.recommendations %}
        <div class=\"section\">
            <h2>Recommendations</h2>
            <ul class=\"recommendations-list\">
                {% for rec in report.recommendations %}
                <li>{{ rec }}</li>
                {% endfor %}
            </ul>
        </div>
        {% endif %}
        {% for table_name, table_data in report.tables.items() %}
        <div class=\"section\">
            <h2>{{ table_name|title }} Details</h2>
            {{ table_data.to_html() | safe }}
        </div>
        {% endfor %}
        <div class=\"footer\">
            <p>Generated by Diriyah Brain AI - Automated Report System</p>
        </div>
    </div>
</body>
</html>
        """
        template_path.write_text(template_html, encoding="utf-8")
        for report_type in ReportType:
            template_variant = self.template_dir / f"{report_type.value}_template.html"
            template_variant.write_text(template_html, encoding="utf-8")

    def _get_report_title(self, report_type: ReportType) -> str:
        mapping = {
            ReportType.DAILY_SUMMARY: "Daily Progress Summary",
            ReportType.WEEKLY_PROGRESS: "Weekly Progress Report",
            ReportType.MONTHLY_OVERVIEW: "Monthly Overview Report",
            ReportType.PROJECT_STATUS: "Project Status Report",
            ReportType.SAFETY_REPORT: "Safety & Compliance Report",
            ReportType.QUALITY_REPORT: "Quality Assurance Report",
            ReportType.FINANCIAL_REPORT: "Financial Performance Report",
            ReportType.EXECUTIVE_SUMMARY: "Executive Summary Report",
        }
        return mapping.get(report_type, "Custom Report")

    def _generate_summary(self, metrics: Dict[str, Any]) -> str:
        completion_rate = metrics.get("completion_rate", 0.0)
        total_alerts = metrics.get("total_alerts", 0)
        critical_alerts = metrics.get("critical_alerts", 0)
        summary_parts: List[str] = []
        if completion_rate >= 80:
            summary_parts.append("Projects are progressing strongly against plan.")
        elif completion_rate >= 60:
            summary_parts.append("Projects remain on track with manageable variances.")
        else:
            summary_parts.append("Projects require intervention to recover productivity.")
        if critical_alerts > 0:
            summary_parts.append(
                f"{critical_alerts} critical alert(s) demand immediate action."
            )
        elif total_alerts > 5:
            summary_parts.append(f"Monitoring {total_alerts} open alerts across the portfolio.")
        else:
            summary_parts.append("Alert levels are within acceptable thresholds.")
        summary_parts.append(
            f"Active projects in focus: {metrics.get('active_projects', 0)}."
        )
        return " ".join(summary_parts)

    def _fig_to_base64(self, fig: Any) -> str:
        buffer = BytesIO()
        fig.savefig(buffer, format="png", dpi=150, bbox_inches="tight")
        buffer.seek(0)
        encoded = base64.b64encode(buffer.read()).decode("utf-8")
        buffer.close()
        return encoded

    def _plotly_to_base64(self, fig: Any) -> str:
        image_bytes = fig.to_image(format="png", width=800, height=600)
        return base64.b64encode(image_bytes).decode("utf-8")


analytics_router = APIRouter(prefix="/analytics", tags=["Analytics"])
reports_router = APIRouter(prefix="/reports", tags=["Reports"])
router = APIRouter()
router.include_router(analytics_router)
router.include_router(reports_router)


@analytics_router.get("/dashboard")
async def get_dashboard_analytics(
    days: int = Query(30, ge=1, le=180, description="Window for analytics in days"),
    current_user: UserContext = Depends(get_current_user),
    db: Any = Depends(_optional_db),
) -> Dict[str, Any]:
    del db  # database not required for stub implementation
    date_from = datetime.utcnow() - timedelta(days=days)
    projects = [
        project
        for project in _SAMPLE_PROJECTS
        if project.user_id == current_user.id
    ]
    tasks = [
        task
        for task in _SAMPLE_TASKS
        if task.user_id == current_user.id and task.created_at >= date_from
    ]
    alerts = [
        alert
        for alert in _SAMPLE_ALERTS
        if alert.user_id == current_user.id and alert.created_at >= date_from
    ]
    total_projects = len(projects)
    active_projects = sum(1 for project in projects if project.status == "active")
    avg_progress = round(
        sum(project.progress for project in projects) / total_projects, 2
    ) if projects else 0.0
    total_tasks = len(tasks)
    completed_tasks = sum(1 for task in tasks if task.completed)
    completion_rate = round(
        (completed_tasks / total_tasks) * 100, 2
    ) if total_tasks else 0.0
    completion_trend: List[Dict[str, Any]] = []
    for offset in range(7):
        day = datetime.utcnow() - timedelta(days=6 - offset)
        day_tasks = [task for task in tasks if task.created_at.date() == day.date()]
        completion_trend.append(
            {
                "date": day.strftime("%Y-%m-%d"),
                "completed": sum(1 for task in day_tasks if task.completed),
                "total": len(day_tasks),
            }
        )
    alert_breakdown = {
        "critical": sum(1 for alert in alerts if alert.severity == "critical"),
        "high": sum(1 for alert in alerts if alert.severity == "high"),
        "medium": sum(1 for alert in alerts if alert.severity == "medium"),
        "low": sum(1 for alert in alerts if alert.severity not in {"critical", "high", "medium"}),
    }
    status_breakdown: Dict[str, int] = {}
    for project in projects:
        status_breakdown[project.status] = status_breakdown.get(project.status, 0) + 1
    return {
        "overview": {
            "total_projects": total_projects,
            "active_projects": active_projects,
            "avg_progress": avg_progress,
            "total_tasks": total_tasks,
            "completed_tasks": completed_tasks,
            "pending_tasks": total_tasks - completed_tasks,
            "completion_rate": completion_rate,
            "total_alerts": len(alerts),
        },
        "trends": {"completion_trend": completion_trend},
        "breakdowns": {
            "alerts": alert_breakdown,
            "project_status": status_breakdown,
        },
        "period": {
            "from": date_from.isoformat(),
            "to": datetime.utcnow().isoformat(),
            "days": days,
        },
    }


@analytics_router.get("/project/{project_id}")
async def get_project_analytics(
    project_id: str,
    current_user: UserContext = Depends(get_current_user),
    db: Any = Depends(_optional_db),
) -> Dict[str, Any]:
    del db
    project = next(
        (
            project
            for project in _SAMPLE_PROJECTS
            if project.id == project_id and project.user_id == current_user.id
        ),
        None,
    )
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    tasks = [task for task in _SAMPLE_TASKS if task.project_id == project_id]
    alerts = [alert for alert in _SAMPLE_ALERTS if alert.project_id == project_id]
    total_tasks = len(tasks)
    completed_tasks = sum(1 for task in tasks if task.completed)
    priority_breakdown = {
        "high": sum(1 for task in tasks if task.priority == "high"),
        "medium": sum(1 for task in tasks if task.priority == "medium"),
        "low": sum(1 for task in tasks if task.priority not in {"high", "medium"}),
    }
    recent_tasks = sorted(tasks, key=lambda item: item.created_at, reverse=True)[:5]
    recent_alerts = sorted(alerts, key=lambda item: item.created_at, reverse=True)[:5]
    return {
        "project": {
            "id": project.id,
            "name": project.name,
            "status": project.status,
            "progress": project.progress,
            "created_at": project.created_at.isoformat(),
        },
        "metrics": {
            "total_tasks": total_tasks,
            "completed_tasks": completed_tasks,
            "pending_tasks": total_tasks - completed_tasks,
            "completion_rate": round((completed_tasks / total_tasks) * 100, 2) if total_tasks else 0.0,
            "total_alerts": len(alerts),
        },
        "breakdowns": {"task_priority": priority_breakdown},
        "recent_activity": {
            "tasks": [
                {
                    "id": task.id,
                    "title": task.title,
                    "status": task.status,
                    "created_at": task.created_at.isoformat(),
                }
                for task in recent_tasks
            ],
            "alerts": [
                {
                    "id": alert.id,
                    "title": alert.title,
                    "severity": alert.severity,
                    "created_at": alert.created_at.isoformat(),
                }
                for alert in recent_alerts
            ],
        },
    }


@analytics_router.get("/performance")
async def get_performance_metrics(
    days: int = Query(30, ge=1, le=180),
    current_user: UserContext = Depends(get_current_user),
    db: Any = Depends(_optional_db),
) -> Dict[str, Any]:
    del db
    date_from = datetime.utcnow() - timedelta(days=days)
    tasks = [
        task
        for task in _SAMPLE_TASKS
        if task.user_id == current_user.id and task.created_at >= date_from
    ]
    alerts = [
        alert
        for alert in _SAMPLE_ALERTS
        if alert.user_id == current_user.id and alert.created_at >= date_from
    ]
    total_tasks = len(tasks)
    completed_tasks = sum(1 for task in tasks if task.completed)
    on_time_tasks = sum(
        1
        for task in tasks
        if task.completed and task.due_date is not None and task.created_at <= task.due_date
    )
    alert_penalty = sum(5 for alert in alerts if alert.severity == "critical")
    productivity_score = 0.0
    if total_tasks:
        completion_factor = (completed_tasks / total_tasks) * 40
        on_time_factor = (on_time_tasks / completed_tasks * 30) if completed_tasks else 0
        alert_factor = max(0, 30 - alert_penalty)
        productivity_score = round(completion_factor + on_time_factor + alert_factor, 2)
    return {
        "kpis": {
            "task_completion_rate": round((completed_tasks / total_tasks) * 100, 2) if total_tasks else 0.0,
            "on_time_completion_rate": round((on_time_tasks / completed_tasks) * 100, 2) if completed_tasks else 0.0,
            "avg_alert_response_time": 2.5,
            "productivity_score": productivity_score,
        },
        "benchmarks": {
            "task_completion_rate": 85,
            "on_time_completion_rate": 90,
            "avg_alert_response_time": 3.0,
            "productivity_score": 75,
        },
        "period": {
            "from": date_from.isoformat(),
            "to": datetime.utcnow().isoformat(),
            "days": days,
        },
    }


@reports_router.post("/generate")
async def generate_report(
    report_type: ReportType,
    report_format: ReportFormat,
    background_tasks: BackgroundTasks,
    project_id: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    current_user: UserContext = Depends(get_current_user),
    db: Any = Depends(_optional_db),
) -> Dict[str, Any]:
    del background_tasks
    generator = AutomatedReportGenerator(db=db, openai_api_key=os.getenv("OPENAI_API_KEY"))
    report_path = await generator.generate_report(
        report_type=report_type,
        report_format=report_format,
        user_id=current_user.id,
        project_id=project_id,
        date_from=date_from,
        date_to=date_to,
    )
    filename = Path(report_path).name
    return {
        "message": "Report generated successfully",
        "report_id": filename,
        "download_url": f"/api/reports/download/{filename}",
        "format": report_format.value,
        "type": report_type.value,
    }


@reports_router.get("/download/{filename}")
async def download_report(filename: str, current_user: UserContext = Depends(get_current_user)) -> FileResponse:
    del current_user
    file_path = Path("./reports") / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Report not found")
    return FileResponse(path=file_path, filename=filename, media_type="application/octet-stream")


@reports_router.get("/list")
async def list_reports(current_user: UserContext = Depends(get_current_user)) -> Dict[str, Any]:
    del current_user
    reports_dir = Path("./reports")
    if not reports_dir.exists():
        return {"reports": []}
    reports: List[Dict[str, Any]] = []
    for file_path in reports_dir.glob("*"):
        if file_path.is_file():
            reports.append(
                {
                    "filename": file_path.name,
                    "size": file_path.stat().st_size,
                    "created_at": datetime.fromtimestamp(file_path.stat().st_ctime).isoformat(),
                    "download_url": f"/api/reports/download/{file_path.name}",
                }
            )
    reports.sort(key=lambda item: item["created_at"], reverse=True)
    return {"reports": reports}


@reports_router.delete("/{filename}")
async def delete_report(filename: str, current_user: UserContext = Depends(get_current_user)) -> Dict[str, str]:
    del current_user
    file_path = Path("./reports") / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Report not found")
    file_path.unlink()
    return {"message": "Report deleted successfully"}
